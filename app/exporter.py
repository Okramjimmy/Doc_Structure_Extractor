"""
Export helpers — JSON, CSV, Excel, Markdown.

All exporters return bytes so they can be streamed directly via
:class:`fastapi.responses.StreamingResponse`.
"""

from __future__ import annotations

import csv
import io
import json
import re
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.schemas import Checklist


# ── JSON ──────────────────────────────────────────────────────────────────────


def to_json(checklist: Checklist, indent: int = 2) -> bytes:
    data = checklist.model_dump(mode="json")
    return json.dumps(data, indent=indent, ensure_ascii=False).encode("utf-8")


# ── Markdown ──────────────────────────────────────────────────────────────────


def to_markdown(checklist: Checklist) -> bytes:
    """
    Generate a structured Markdown checklist document.

    Layout:
      # Document title
      > metadata block
      ## Table of Contents
      ## Section Name
      | No. | Item | Response | Remarks |
      ...
      ## Summary
    """
    import re as _re
    from collections import Counter

    lines: list[str] = []

    # ── Title ─────────────────────────────────────────────────────────────────
    doc_stem = checklist.document_name.rsplit(".", 1)[0]
    # Strip leading UUID prefix added by the upload handler (e.g. "uuid4_filename")
    doc_stem = _re.sub(r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}_", "", doc_stem)
    lines += [
        f"# {doc_stem} — Quality Checklist",
        "",
        f"> **Document:** {checklist.document_name}  ",
        f"> **Extracted:** {checklist.extracted_at.strftime('%Y-%m-%d %H:%M UTC')}  ",
        f"> **Total Items:** {checklist.total_questions}  ",
        f"> **Sections:** {len(checklist.sections)}  ",
        "",
    ]

    # ── Table of Contents ─────────────────────────────────────────────────────
    lines += ["## Table of Contents", ""]
    for i, section in enumerate(checklist.sections, 1):
        # GitHub-style anchor: lowercase, spaces → hyphens, strip non-alphanum
        anchor = re.sub(r"[^a-z0-9\s-]", "", section.lower()).strip().replace(" ", "-")
        anchor = re.sub(r"-+", "-", anchor)
        lines.append(f"{i}. [{section}](#{anchor})")
    lines += ["", "---", ""]

    # ── Per-section tables ────────────────────────────────────────────────────
    # Group questions by section preserving order
    sections_map: dict[str, list] = {}
    for q in checklist.questions:
        sections_map.setdefault(q.section, []).append(q)

    for section, qs in sections_map.items():
        lines += [f"## {section}", ""]

        # Section path breadcrumb (skip if it's just the section itself)
        if qs and len(qs[0].section_path) > 1:
            breadcrumb = " › ".join(qs[0].section_path[:-1])
            lines += [f"*{breadcrumb}*", ""]

        # Table header
        lines += [
            "| No. | Code | Checklist Item | Response | Remarks |",
            "|-----|------|---------------|----------|---------|",
        ]

        for row_num, q in enumerate(qs, 1):
            # Escape pipe characters inside cell text
            item_text = q.question.replace("|", "\\|")
            response_hint = {
                "yes_no_na": "Yes / No / N/A",
                "yes_no":    "Yes / No",
                "text":      "_(text)_",
                "numeric":   "_(number)_",
            }.get(q.response_type.value, "")

            lines.append(
                f"| {row_num} | `{q.question_code}` | {item_text} | {response_hint} |  |"
            )

        lines += ["", ""]

    # ── Summary ───────────────────────────────────────────────────────────────
    lines += ["---", "", "## Summary", ""]
    lines += [
        "| Section | Items |",
        "|---------|-------|",
    ]
    counts = Counter(q.section for q in checklist.questions)
    for section in checklist.sections:
        lines.append(f"| {section} | {counts[section]} |")

    lines += [
        "",
        f"**Total: {checklist.total_questions} checklist items across "
        f"{len(checklist.sections)} sections.**",
        "",
    ]

    return "\n".join(lines).encode("utf-8")




# ── CSV ───────────────────────────────────────────────────────────────────────

_CSV_HEADERS = [
    "question_code",
    "section",
    "section_path",
    "question",
    "response_type",
    "response",  # blank — for the user to fill in
    "remarks",
]


def to_csv(checklist: Checklist) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=_CSV_HEADERS, lineterminator="\r\n")
    writer.writeheader()
    for q in checklist.questions:
        writer.writerow(
            {
                "question_code": q.question_code,
                "section": q.section,
                "section_path": " > ".join(q.section_path),
                "question": q.question,
                "response_type": q.response_type.value,
                "response": "",
                "remarks": "",
            }
        )
    return buf.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility


# ── Excel ─────────────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1E293B")
_SECTION_FILL = PatternFill("solid", fgColor="334155")
_ALT_FILL = PatternFill("solid", fgColor="F8FAFC")

_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
_SECTION_FONT = Font(bold=True, color="94A3B8", name="Calibri", size=10, italic=True)
_BODY_FONT = Font(name="Calibri", size=10)
_CODE_FONT = Font(name="Consolas", size=10, color="6366F1")


def to_excel(checklist: Checklist) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Checklist"

    # ── Cover info ────────────────────────────────────────────────────────────
    ws.append(["Document", checklist.document_name])
    ws.append(["Extracted at", checklist.extracted_at.strftime("%Y-%m-%d %H:%M UTC")])
    ws.append(["Total Questions", checklist.total_questions])
    ws.append([])  # spacer

    # ── Header row ────────────────────────────────────────────────────────────
    headers = ["Code", "Section", "Section Path", "Question", "Response Type", "Response", "Remarks"]
    ws.append(headers)
    header_row = ws.max_row
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Data rows ─────────────────────────────────────────────────────────────
    prev_section = None
    for i, q in enumerate(checklist.questions):
        row_num = ws.max_row + 1
        is_alt = (i % 2 == 0)

        # Section separator
        if q.section != prev_section:
            ws.append([f"── {q.section} ──", "", "", "", "", "", ""])
            sep_row = ws.max_row
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=sep_row, column=col_idx)
                cell.fill = _SECTION_FILL
                cell.font = _SECTION_FONT
                cell.alignment = Alignment(vertical="center")
            ws.merge_cells(f"A{sep_row}:G{sep_row}")
            prev_section = q.section

        ws.append(
            [
                q.question_code,
                q.section,
                " > ".join(q.section_path),
                q.question,
                q.response_type.value,
                "",   # Response — user fills in
                "",   # Remarks
            ]
        )
        data_row = ws.max_row
        fill = _ALT_FILL if is_alt else None
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=data_row, column=col_idx)
            if fill:
                cell.fill = fill
            cell.font = _CODE_FONT if col_idx == 1 else _BODY_FONT
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=(col_idx == 4),  # wrap question text
            )

        # Response dropdown for yes_no_na
        if q.response_type.value == "yes_no_na":
            from openpyxl.worksheet.datavalidation import DataValidation
            dv = DataValidation(
                type="list",
                formula1='"Yes,No,N/A"',
                showDropDown=False,
            )
            ws.add_data_validation(dv)
            dv.add(ws.cell(row=data_row, column=6))

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [10, 22, 35, 70, 15, 12, 20]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Freeze panes ─────────────────────────────────────────────────────────
    ws.freeze_panes = f"A{header_row + 1}"

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Section", "Question Count"])
    from collections import Counter
    counts = Counter(q.section for q in checklist.questions)
    for section, count in counts.most_common():
        ws2.append([section, count])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
